"""
Excel 보고서 템플릿 채우기 스크립트

Usage:
    py fill_excel_template.py <data_json_path>

data_json_path: consult 스킬이 생성하는 임시 JSON 파일 경로
                (output/temp_[session_id].json)

JSON 스키마 → consult/SKILL.md STEP 7 참조
"""
import sys
import json
import shutil
import os
import math
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── 경로 설정 ─────────────────────────────────────────────────
_HERE         = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(_HERE, "컨설팅결과_보고서_템플릿.xlsx")
KR_SHEET      = "컨설팅 결과 보고서"
EN_SHEET      = "Consulting Report (EN)"


# ── 플레이스홀더 교체 ──────────────────────────────────────────
def _fill_cell(cell, data: dict):
    """셀 값 내 {{KEY}} 를 data[KEY]로 일괄 교체. 누락된 키는 빈 문자열로 처리."""
    if not isinstance(cell.value, str) or "{{" not in cell.value:
        return
    import re
    cell.value = re.sub(
        r'\{\{(\w+)\}\}',
        lambda m: str(data.get(m.group(1), "")),
        cell.value
    )


def fill_sheet(ws, data: dict):
    """워크시트 전체 셀의 플레이스홀더를 채운다."""
    for row in ws.iter_rows():
        for cell in row:
            _fill_cell(cell, data)


# ── Revision 데이터 평탄화 ────────────────────────────────────
def _flatten_revisions(data: dict, revisions: list, rev_total: int,
                        rev_light: int, rev_full: int):
    """
    revisions 목록을 REV_1_NO, REV_1_TYPE … REV_3_REASON 형태로 평탄화.
    템플릿 행(3개) 초과분은 첫 3개만 사용.
    """
    for i in range(1, 4):
        if i <= len(revisions):
            rev = revisions[i - 1]
            data[f"REV_{i}_NO"]     = rev.get("NO", "")
            data[f"REV_{i}_TYPE"]   = rev.get("TYPE", "")
            data[f"REV_{i}_PREV"]   = rev.get("PREV", "")
            data[f"REV_{i}_NEW"]    = rev.get("NEW", "")
            data[f"REV_{i}_REASON"] = rev.get("REASON", "")
        else:
            data[f"REV_{i}_NO"]     = ""
            data[f"REV_{i}_TYPE"]   = ""
            data[f"REV_{i}_PREV"]   = ""
            data[f"REV_{i}_NEW"]    = ""
            data[f"REV_{i}_REASON"] = ""

    data["REV_TOTAL"] = rev_total
    data["REV_LIGHT"] = rev_light
    data["REV_FULL"]  = rev_full


# ── 자동 행 높이 조절 ─────────────────────────────────────────
def _estimate_lines(text: str, col_width: float) -> int:
    """텍스트가 주어진 열 너비(문자 단위)에서 몇 줄을 차지하는지 추정."""
    if not text or not text.strip():
        return 1
    # 열 너비 1unit ≈ 약 1.6 문자 (맑은 고딕 10pt 기준 경험값)
    chars_per_line = max(1, int(col_width * 1.6))
    total = 0
    for line in text.split("\n"):
        total += max(1, math.ceil(len(line) / chars_per_line))
    return max(1, total)


def remove_italic(ws):
    """시트 전체 셀에서 기울임꼴을 제거한다."""
    from openpyxl.styles import Font
    for row in ws.iter_rows():
        for cell in row:
            if cell.font and cell.font.italic:
                f = cell.font
                cell.font = Font(
                    name=f.name, bold=f.bold, size=f.size,
                    color=f.color, italic=False,
                    underline=f.underline, strike=f.strike,
                )


def auto_adjust_row_heights(ws, line_height_pt: float = 14.5, padding_pt: float = 6):
    """
    fill_sheet() 이후 호출. 각 행의 실제 내용 기준으로 높이를 재계산한다.
    - 병합 셀은 병합된 전체 너비로 계산
    - 섹션 타이틀 행(배경색이 진한 행)은 최소 24pt 유지
    """
    # 열 너비 수집
    col_widths: dict[int, float] = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        idx = openpyxl.utils.column_index_from_string(col_letter)
        col_widths[idx] = col_dim.width if col_dim.width else 8.0

    # 병합 셀 → 합산 너비 매핑  {(row, col): total_width}
    merge_widths: dict[tuple, float] = {}
    for rng in ws.merged_cells.ranges:
        total_w = sum(col_widths.get(c, 8.0)
                      for c in range(rng.min_col, rng.max_col + 1))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_widths[(r, c)] = total_w

    for row_idx in range(1, ws.max_row + 1):
        max_lines = 1
        for cell in ws[row_idx]:
            if not isinstance(cell.value, str) or not cell.value.strip():
                continue
            col_w = merge_widths.get(
                (row_idx, cell.column),
                col_widths.get(cell.column, 8.0)
            )
            lines = _estimate_lines(cell.value, col_w)
            max_lines = max(max_lines, lines)

        new_height = max(18, max_lines * line_height_pt + padding_pt)
        ws.row_dimensions[row_idx].height = new_height


# ── 동적 Revision 행 추가 (4건 초과 시) ──────────────────────
def _insert_extra_revision_rows(ws, revisions: list):
    """
    재컨설팅 이력이 3건 초과일 때 REV_3 행 바로 아래에 추가 행을 삽입한다.
    추가된 행에는 REV_N_* placeholder를 직접 기입하지 않고
    빈 행으로만 삽입 (3건 초과분은 별도 처리 없이 빈칸 허용).
    """
    if len(revisions) <= 3:
        return

    # REV_3_NO 셀이 있는 행 탐색
    rev3_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "{{REV_3_NO}}" in cell.value:
                rev3_row_idx = cell.row
                break
        if rev3_row_idx:
            break

    if rev3_row_idx is None:
        return

    extra = revisions[3:]
    for offset, rev in enumerate(extra):
        insert_at = rev3_row_idx + 1 + offset
        ws.insert_rows(insert_at)
        ws.cell(row=insert_at, column=1, value=rev.get("NO", ""))
        ws.cell(row=insert_at, column=2, value=rev.get("TYPE", ""))
        ws.cell(row=insert_at, column=3, value=rev.get("PREV", ""))
        ws.cell(row=insert_at, column=4, value=rev.get("NEW", ""))
        ws.cell(row=insert_at, column=5, value=rev.get("REASON", ""))


# ── 메인 ──────────────────────────────────────────────────────
def main():
    if len(sys.argv) < 2:
        print("[ERROR] Usage: py fill_excel_template.py <data_json_path>")
        sys.exit(1)

    json_path = sys.argv[1]

    if not os.path.exists(json_path):
        print(f"[ERROR] JSON 파일 없음: {json_path}")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        payload = json.load(f)

    output_language = payload.get("output_language", "ko")
    output_path     = payload.get("output_path", "")
    revisions       = payload.get("revisions", [])
    rev_total       = payload.get("rev_total", 0)
    rev_light       = payload.get("rev_light", 0)
    rev_full        = payload.get("rev_full", 0)

    if not output_path:
        print("[ERROR] output_path가 지정되지 않았습니다.")
        sys.exit(1)

    # 출력 디렉토리 생성 & 템플릿 복사
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    if not os.path.exists(TEMPLATE_PATH):
        print(f"[ERROR] 템플릿 파일 없음: {TEMPLATE_PATH}")
        sys.exit(1)

    shutil.copy2(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)

    # ── KR 시트 채우기 ────────────────────────────────────────
    if output_language in ("ko", "en+ko") and "ko" in payload:
        ws_kr = wb[KR_SHEET]
        data_kr = dict(payload["ko"])
        _flatten_revisions(data_kr, revisions, rev_total, rev_light, rev_full)
        _insert_extra_revision_rows(ws_kr, revisions)
        fill_sheet(ws_kr, data_kr)
        remove_italic(ws_kr)
        auto_adjust_row_heights(ws_kr)

    # ── EN 시트 채우기 ────────────────────────────────────────
    if output_language in ("en", "en+ko") and "en" in payload:
        ws_en = wb[EN_SHEET]
        data_en = dict(payload["en"])
        _flatten_revisions(data_en, revisions, rev_total, rev_light, rev_full)
        _insert_extra_revision_rows(ws_en, revisions)
        fill_sheet(ws_en, data_en)
        remove_italic(ws_en)
        auto_adjust_row_heights(ws_en)

    wb.save(output_path)
    print(f"[OK] Excel 보고서 저장 완료: {output_path}")


if __name__ == "__main__":
    main()
