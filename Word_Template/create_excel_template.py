"""
컨설팅 결과 보고서 Excel 템플릿 생성 스크립트
실행: py create_excel_template.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "C:/Users/sy.ahn/OneDrive - Bosch Group/AI Development/AI_Consulting/Word_Template/컨설팅결과_보고서_템플릿.xlsx"

# ── 색상 팔레트 ──────────────────────────────────────────────
C_BRAND_DARK  = "1A3A5C"   # 진남색 (헤더 배경)
C_BRAND_MID   = "2E6DA4"   # 중간 파랑 (섹션 타이틀)
C_BRAND_LIGHT = "D6E4F0"   # 연한 파랑 (레이블 배경)
C_GREEN_BG    = "E8F5E9"   # 연초록 (권장안)
C_GREEN_HDR   = "2E7D32"   # 진초록 (권장 배지)
C_YELLOW_BG   = "FFFDE7"   # 연노랑 (검토필요안)
C_YELLOW_HDR  = "F57F17"   # 진노랑 (검토필요 배지)
C_RED_BG      = "FFEBEE"   # 연빨강 (비추천안)
C_RED_HDR     = "C62828"   # 진빨강 (비추천 배지)
C_GRAY_BG     = "F5F5F5"   # 연회색 (교대 행)
C_GRAY_HDR    = "616161"   # 회색 (보조 헤더)
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"
C_TEXT_MUTED  = "757575"
C_PURPLE_BG   = "F3E5F5"   # AI 분석 섹션
C_PURPLE_HDR  = "6A1B9A"

# ── 헬퍼 함수 ────────────────────────────────────────────────
def _font(bold=False, size=10, color=C_BLACK, italic=False):
    return Font(name="맑은 고딕", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color="BDBDBD"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(ws, row, col, value,
         bold=False, size=10, font_color=C_BLACK, italic=False,
         fill=None, h="left", v="center", wrap=True, border=True, border_style="thin"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=size, color=font_color, italic=italic)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _border(style=border_style)
    return cell

def _merge(ws, r1, c1, r2, c2, value="",
           bold=False, size=10, font_color=C_BLACK, italic=False,
           fill=None, h="left", v="center", wrap=True, border=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font      = _font(bold=bold, size=size, color=font_color, italic=italic)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _border()
    return cell

def _section_title(ws, row, title, color=C_BRAND_MID):
    """6열 병합 섹션 타이틀"""
    _merge(ws, row, 1, row, 6, value=title,
           bold=True, size=12, font_color=C_WHITE, fill=color, h="left", border=False)
    ws.row_dimensions[row].height = 24

def _label(ws, row, col, text):
    _set(ws, row, col, text, bold=True, font_color=C_BRAND_DARK,
         fill=C_BRAND_LIGHT, h="center")

def _value(ws, row, col, placeholder, fill=C_WHITE, colspan=1, ws_ref=None):
    if colspan > 1:
        _merge(ws, row, col, row, col + colspan - 1, value=placeholder,
               font_color="555555", italic=False, fill=fill)
    else:
        _set(ws, row, col, placeholder, font_color="555555", italic=False, fill=fill)

# ─────────────────────────────────────────────────────────────
def build_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "컨설팅 결과 보고서"

    # 열 너비 설정
    col_widths = {1: 18, 2: 28, 3: 18, 4: 28, 5: 18, 6: 28}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # 기본 행 높이
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1

    row = 1

    # ══════════════════════════════════════════════════
    # TITLE
    # ══════════════════════════════════════════════════
    ws.row_dimensions[row].height = 14
    row += 1

    _merge(ws, row, 1, row+1, 6,
           value="MS 업무 자동화  컨설팅 결과 보고서",
           bold=True, size=18, font_color=C_WHITE,
           fill=C_BRAND_DARK, h="center", v="center")
    ws.row_dimensions[row].height = 36
    ws.row_dimensions[row+1].height = 20
    row += 2

    ws.row_dimensions[row].height = 6
    row += 1

    # ══════════════════════════════════════════════════
    # 0. 메타 정보
    # ══════════════════════════════════════════════════
    _section_title(ws, row, "  세션 정보", C_GRAY_HDR)
    row += 1

    meta = [
        ("세션 ID",  "{{SESSION_ID}}",    "모드",      "{{MODE}}"),
        ("생성일시",  "{{CREATED_AT}}",   "업무 도메인", "{{DOMAIN}}"),
        ("MS 지원",  "{{MS_VERIFY}}",     "파싱 신뢰도", "{{PARSE_CONFIDENCE}}%"),
    ]
    for lbl1, val1, lbl2, val2 in meta:
        ws.row_dimensions[row].height = 20
        _label(ws, row, 1, lbl1);  _value(ws, row, 2, val1)
        _label(ws, row, 3, lbl2);  _value(ws, row, 4, val2)
        _set(ws, row, 5, ""); _set(ws, row, 6, "")  # 빈 열
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 1. 요구사항 요약
    # ══════════════════════════════════════════════════
    _section_title(ws, row, "  1.  요구사항 요약")
    row += 1

    req_fields = [
        ("업무 도메인",   "{{REQ_DOMAIN}}",     "프로세스 유형", "{{REQ_PROCESS_TYPE}}"),
        ("자동화 대상",   "{{REQ_TARGET}}",     "발송 규모",    "{{REQ_SCALE}}"),
        ("현재 도구",    "{{REQ_CURRENT_TOOL}}","첨부파일",     "{{REQ_ATTACHMENT}}"),
        ("외부 시스템",  "{{REQ_EXT_SYSTEM}}",  "이력 저장",    "{{REQ_HISTORY_SAVE}}"),
        ("제약 조건",    "{{REQ_CONSTRAINTS}}", "",            ""),
    ]
    for lbl1, val1, lbl2, val2 in req_fields:
        ws.row_dimensions[row].height = 22
        _label(ws, row, 1, lbl1)
        if lbl2:
            _value(ws, row, 2, val1)
            _label(ws, row, 3, lbl2)
            _value(ws, row, 4, val2)
            _set(ws, row, 5, ""); _set(ws, row, 6, "")
        else:
            _value(ws, row, 2, val1, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 2. 사용자용 솔루션 안내
    # ══════════════════════════════════════════════════
    _section_title(ws, row, "  2.  사용자용 솔루션 안내")
    row += 1

    def _solution_user(row, badge_color, badge_text, bg, prefix):
        ws.row_dimensions[row].height = 22
        _merge(ws, row, 1, row, 6,
               value=f"  {badge_text}   {{{{SOL_{prefix}_NAME}}}}",
               bold=True, size=11, font_color=C_WHITE, fill=badge_color, border=False)
        row += 1

        fields_u = [
            ("솔루션명",   f"{{{{SOL_{prefix}_NAME}}}}"),
            ("추천 이유",  f"{{{{SOL_{prefix}_REASON}}}}"),
            ("구현 흐름",  f"{{{{SOL_{prefix}_FLOW}}}}"),
            ("준비사항",   f"{{{{SOL_{prefix}_PREREQ}}}}"),
            ("주의사항",   f"{{{{SOL_{prefix}_CAUTION}}}}"),
        ]
        for lbl, val in fields_u:
            ws.row_dimensions[row].height = 40 if lbl in ("구현 흐름","준비사항","주의사항") else 22
            _label(ws, row, 1, lbl)
            _value(ws, row, 2, val, fill=bg, colspan=5)
            row += 1
        return row

    row = _solution_user(row, C_GREEN_HDR,  "[ 권장 ]",      C_GREEN_BG,  "REC")
    row += 1
    row = _solution_user(row, C_YELLOW_HDR, "[ 검토 필요 ]", C_YELLOW_BG, "REV")
    row += 1

    # 비추천 (간소)
    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, row, 6,
           value="  [ 비추천 ]   {{SOL_NOT_NAME}}",
           bold=True, size=11, font_color=C_WHITE, fill=C_RED_HDR, border=False)
    row += 1
    for lbl, val in [("솔루션명", "{{SOL_NOT_NAME}}"), ("비추천 이유", "{{SOL_NOT_REASON}}")]:
        ws.row_dimensions[row].height = 22
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_RED_BG, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 3. 개발자용 기술 명세
    # ══════════════════════════════════════════════════
    _section_title(ws, row, "  3.  개발자용 기술 명세")
    row += 1

    # 권장안 기술 명세
    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, row, 6,
           value="  [ 권장 ]   {{SOL_REC_NAME}}",
           bold=True, size=11, font_color=C_WHITE, fill=C_GREEN_HDR, border=False)
    row += 1

    dev_fields = [
        ("솔루션명 / ID",  "{{SOL_REC_NAME}}  |  ID: {{SOL_REC_ID}}"),
        ("적용 이유",     "{{SOL_REC_TECH_REASON}}"),
        ("구현 개요",     "{{SOL_REC_IMPL_OVERVIEW}}"),
        ("전제조건",      "{{SOL_REC_PREREQ_TECH}}"),
        ("한계점",       "{{SOL_REC_LIMITS}}"),
        ("고려사항",      "{{SOL_REC_CONSIDERATIONS}}"),
    ]
    for lbl, val in dev_fields:
        h = 50 if lbl in ("구현 개요","전제조건","한계점","고려사항") else 22
        ws.row_dimensions[row].height = h
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_GREEN_BG, colspan=5)
        row += 1

    # 리스크 테이블
    ws.row_dimensions[row].height = 20
    _label(ws, row, 1, "리스크 상세")
    _set(ws, row, 2, ""); _set(ws, row, 3, ""); _set(ws, row, 4, ""); _set(ws, row, 5, ""); _set(ws, row, 6, "")
    row += 1

    risk_header = ["유형", "영향도", "발생 조건", "대응 방안", "", ""]
    for c_idx, h_val in enumerate(risk_header, 1):
        if h_val:
            _set(ws, row, c_idx, h_val, bold=True, fill=C_BRAND_LIGHT,
                 font_color=C_BRAND_DARK, h="center")
        else:
            _set(ws, row, c_idx, "")
    row += 1

    risk_rows = [
        ("보안",      "높음",   "{{RISK_SEC_CONDITION}}",   "{{RISK_SEC_ACTION}}"),
        ("라이선스",  "중간",   "{{RISK_LIC_CONDITION}}",   "{{RISK_LIC_ACTION}}"),
        ("운영",      "중간",   "{{RISK_OPS_CONDITION}}",   "{{RISK_OPS_ACTION}}"),
    ]
    for r_type, r_level, r_cond, r_act in risk_rows:
        ws.row_dimensions[row].height = 30
        _set(ws, row, 1, r_type,  bold=True, fill=C_GRAY_BG, h="center")
        _set(ws, row, 2, r_level, fill=C_GRAY_BG, h="center")
        _merge(ws, row, 3, row, 4, value=r_cond, font_color="555555", italic=False, fill=C_WHITE)
        _merge(ws, row, 5, row, 6, value=r_act,  font_color="555555", italic=False, fill=C_WHITE)
        row += 1

    row += 1

    # 검토필요 / 비추천 기술 명세 (간소)
    for badge_color, badge_text, bg, prefix in [
        (C_YELLOW_HDR, "[ 검토 필요 ]", C_YELLOW_BG, "REV"),
        (C_RED_HDR,    "[ 비추천 ]",   C_RED_BG,    "NOT"),
    ]:
        ws.row_dimensions[row].height = 22
        _merge(ws, row, 1, row, 6,
               value=f"  {badge_text}   {{{{SOL_{prefix}_NAME}}}}",
               bold=True, size=11, font_color=C_WHITE, fill=badge_color, border=False)
        row += 1
        for lbl, val in [
            ("솔루션명 / ID", f"{{{{SOL_{prefix}_NAME}}}}  |  ID: {{{{SOL_{prefix}_ID}}}}"),
            ("적용 이유",    f"{{{{SOL_{prefix}_TECH_REASON}}}}"),
            ("판정 근거",    f"{{{{SOL_{prefix}_JUDGMENT}}}}"),
        ]:
            ws.row_dimensions[row].height = 30
            _label(ws, row, 1, lbl)
            _value(ws, row, 2, val, fill=bg, colspan=5)
            row += 1
        row += 1

    # ══════════════════════════════════════════════════
    # 4. AI 정밀 분석
    # ══════════════════════════════════════════════════
    _section_title(ws, row, "  4.  AI 정밀 분석", C_PURPLE_HDR)
    row += 1

    ai_fields = [
        ("분석 개요",        "{{AI_OVERVIEW}}"),
        ("핵심 발견사항",    "{{AI_FINDINGS}}"),
        ("최적화 권고",      "{{AI_RECOMMENDATIONS}}"),
        ("추가 리스크/기회", "{{AI_EXTRA_RISK_OPP}}"),
    ]
    for lbl, val in ai_fields:
        ws.row_dimensions[row].height = 50
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_PURPLE_BG, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 부록 A. 재컨설팅 이력
    # ══════════════════════════════════════════════════
    _section_title(ws, row, "  부록 A.  재컨설팅 이력", C_GRAY_HDR)
    row += 1

    rev_headers = ["Rev", "타입", "이전 권고", "새 권고", "변경 이유", ""]
    for c_idx, h_val in enumerate(rev_headers, 1):
        _set(ws, row, c_idx, h_val, bold=True, fill=C_BRAND_LIGHT,
             font_color=C_BRAND_DARK, h="center")
    row += 1

    # 이력 행 3개 (동적 추가를 위한 예시 행)
    for i in range(1, 4):
        ws.row_dimensions[row].height = 22
        _set(ws, row, 1, f"{{{{REV_{i}_NO}}}}", fill=C_GRAY_BG, h="center")
        _set(ws, row, 2, f"{{{{REV_{i}_TYPE}}}}", fill=C_GRAY_BG, h="center")
        _set(ws, row, 3, f"{{{{REV_{i}_PREV}}}}", font_color="555555", italic=False)
        _set(ws, row, 4, f"{{{{REV_{i}_NEW}}}}", font_color="555555", italic=False)
        _merge(ws, row, 5, row, 6, value=f"{{{{REV_{i}_REASON}}}}",
               font_color="555555", italic=False)
        row += 1

    # 요약 행
    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, row, 6,
           value="총 재컨설팅 횟수: {{REV_TOTAL}}회  (경량: {{REV_LIGHT}}회  |  전체: {{REV_FULL}}회)",
           bold=True, fill=C_BRAND_LIGHT, font_color=C_BRAND_DARK, h="right")
    row += 1

    row += 1

    # ── 하단 여백 & 범례 ─────────────────────────────
    _merge(ws, row, 1, row, 6,
           value="※  {{변수명}} 형식의 셀을 실제 데이터로 교체하여 사용합니다.",
           italic=False, font_color=C_TEXT_MUTED, fill=C_GRAY_BG,
           h="center", border=False)

    wb.save(OUTPUT_PATH)
    print(f"[OK] 템플릿 저장 완료: {OUTPUT_PATH}")

if __name__ == "__main__":
    build_template()
