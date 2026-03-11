"""
영문 시트를 기존 Excel 템플릿에 추가하는 스크립트
실행: py add_english_sheet.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILE_PATH = "C:/Users/sy.ahn/OneDrive - Bosch Group/AI Development/AI_Consulting/Word_Template/컨설팅결과_보고서_템플릿.xlsx"
EN_SHEET  = "Consulting Report (EN)"

# ── Color palette (same as KR sheet) ────────────────────────
C_BRAND_DARK  = "1A3A5C"
C_BRAND_MID   = "2E6DA4"
C_BRAND_LIGHT = "D6E4F0"
C_GREEN_BG    = "E8F5E9"
C_GREEN_HDR   = "2E7D32"
C_YELLOW_BG   = "FFFDE7"
C_YELLOW_HDR  = "F57F17"
C_RED_BG      = "FFEBEE"
C_RED_HDR     = "C62828"
C_GRAY_BG     = "F5F5F5"
C_GRAY_HDR    = "616161"
C_WHITE       = "FFFFFF"
C_BLACK       = "000000"
C_TEXT_MUTED  = "757575"
C_PURPLE_BG   = "F3E5F5"
C_PURPLE_HDR  = "6A1B9A"
C_TEAL_BG     = "E0F2F1"
C_TEAL_HDR    = "00695C"

# ── Helpers ──────────────────────────────────────────────────
def _font(bold=False, size=10, color=C_BLACK, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border(style="thin", color="BDBDBD"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set(ws, row, col, value,
         bold=False, size=10, font_color=C_BLACK, italic=False,
         fill=None, h="left", v="center", wrap=True, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = _font(bold=bold, size=size, color=font_color, italic=italic)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _border()
    return cell

def _merge(ws, r1, c1, r2, c2, value="",
           bold=False, size=10, font_color=C_BLACK, italic=False,
           fill=None, h="left", v="center", wrap=True, border=True):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1, value=value)
    cell.font      = _font(bold=bold, size=size, color=font_color, italic=italic)
    cell.alignment = _align(h=h, v=v, wrap=wrap)
    if fill:
        cell.fill = _fill(fill)
    if border:
        cell.border = _border()
    return cell

def _section(ws, row, title, color=C_BRAND_MID):
    _merge(ws, row, 1, row, 6, value=title,
           bold=True, size=12, font_color=C_WHITE, fill=color, h="left", border=False)
    ws.row_dimensions[row].height = 24

def _label(ws, row, col, text):
    _set(ws, row, col, text, bold=True, font_color=C_BRAND_DARK,
         fill=C_BRAND_LIGHT, h="center")

def _value(ws, row, col, placeholder, fill=C_WHITE, colspan=1):
    if colspan > 1:
        _merge(ws, row, col, row, col + colspan - 1,
               value=placeholder, font_color="555555", italic=False, fill=fill)
    else:
        _set(ws, row, col, placeholder, font_color="555555", italic=False, fill=fill)

# ── Main builder ─────────────────────────────────────────────
def build_english_sheet(wb):
    # Remove existing EN sheet if present
    if EN_SHEET in wb.sheetnames:
        del wb[EN_SHEET]

    ws = wb.create_sheet(title=EN_SHEET)

    # Column widths
    for c, w in {1: 20, 2: 30, 3: 20, 4: 30, 5: 18, 6: 28}.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1

    row = 1

    # ─── Top padding ─────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    row += 1

    # ══════════════════════════════════════════════════
    # TITLE
    # ══════════════════════════════════════════════════
    _merge(ws, row, 1, row + 1, 6,
           value="MS Work Automation  Consulting Report",
           bold=True, size=18, font_color=C_WHITE,
           fill=C_BRAND_DARK, h="center", v="center")
    ws.row_dimensions[row].height = 36
    ws.row_dimensions[row + 1].height = 20
    row += 2

    ws.row_dimensions[row].height = 6
    row += 1

    # ══════════════════════════════════════════════════
    # Session Info
    # ══════════════════════════════════════════════════
    _section(ws, row, "  Session Info", C_GRAY_HDR)
    row += 1

    meta = [
        ("Session ID",        "{{SESSION_ID}}",          "Mode",             "{{MODE}}"),
        ("Created At",        "{{CREATED_AT}}",          "Business Domain",  "{{DOMAIN}}"),
        ("MS Support",        "{{MS_VERIFY}}",           "Parse Confidence", "{{PARSE_CONFIDENCE}}%"),
    ]
    for lbl1, val1, lbl2, val2 in meta:
        ws.row_dimensions[row].height = 20
        _label(ws, row, 1, lbl1);  _value(ws, row, 2, val1)
        _label(ws, row, 3, lbl2);  _value(ws, row, 4, val2)
        _set(ws, row, 5, "");      _set(ws, row, 6, "")
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 1. Requirements Summary
    # ══════════════════════════════════════════════════
    _section(ws, row, "  1.  Requirements Summary")
    row += 1

    req_fields = [
        ("Business Domain",  "{{REQ_DOMAIN}}",       "Process Type",      "{{REQ_PROCESS_TYPE}}"),
        ("Automation Target","{{REQ_TARGET}}",        "Send Volume",       "{{REQ_SCALE}}"),
        ("Current Tools",    "{{REQ_CURRENT_TOOL}}", "Attachments",       "{{REQ_ATTACHMENT}}"),
        ("External Systems", "{{REQ_EXT_SYSTEM}}",   "History Logging",   "{{REQ_HISTORY_SAVE}}"),
        ("Constraints",      "{{REQ_CONSTRAINTS}}",  "",                  ""),
    ]
    for lbl1, val1, lbl2, val2 in req_fields:
        ws.row_dimensions[row].height = 22
        _label(ws, row, 1, lbl1)
        if lbl2:
            _value(ws, row, 2, val1)
            _label(ws, row, 3, lbl2)
            _value(ws, row, 4, val2)
            _set(ws, row, 5, ""); _set(ws, row, 6, "")
        else:
            _value(ws, row, 2, val1, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 2. User-Facing Solution Guide
    # ══════════════════════════════════════════════════
    _section(ws, row, "  2.  User-Facing Solution Guide")
    row += 1

    def _solution_user(row, badge_color, badge_text, bg, prefix,
                       fields):
        ws.row_dimensions[row].height = 22
        _merge(ws, row, 1, row, 6,
               value=f"  {badge_text}   {{{{SOL_{prefix}_NAME}}}}",
               bold=True, size=11, font_color=C_WHITE, fill=badge_color, border=False)
        row += 1
        for lbl, val in fields:
            ws.row_dimensions[row].height = (
                40 if lbl in ("Implementation Flow", "Prerequisites", "Cautions") else 22
            )
            _label(ws, row, 1, lbl)
            _value(ws, row, 2, val, fill=bg, colspan=5)
            row += 1
        return row

    rec_fields = [
        ("Solution Name",        "{{SOL_REC_NAME}}"),
        ("Recommendation Reason","{{SOL_REC_REASON}}"),
        ("Implementation Flow",  "{{SOL_REC_FLOW}}"),
        ("Prerequisites",        "{{SOL_REC_PREREQ}}"),
        ("Cautions",             "{{SOL_REC_CAUTION}}"),
    ]
    row = _solution_user(row, C_GREEN_HDR,  "[ Recommended ]",    C_GREEN_BG,  "REC", rec_fields)
    row += 1

    rev_fields = [
        ("Solution Name",       "{{SOL_REV_NAME}}"),
        ("Reason / Limitation", "{{SOL_REV_REASON}}"),
        ("Judgment Basis",      "{{SOL_REV_JUDGMENT}}"),
    ]
    row = _solution_user(row, C_YELLOW_HDR, "[ Needs Review ]",   C_YELLOW_BG, "REV", rev_fields)
    row += 1

    # Not Recommended (compact)
    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, row, 6,
           value="  [ Not Recommended ]   {{SOL_NOT_NAME}}",
           bold=True, size=11, font_color=C_WHITE, fill=C_RED_HDR, border=False)
    row += 1
    for lbl, val in [("Solution Name", "{{SOL_NOT_NAME}}"),
                     ("Reason",        "{{SOL_NOT_REASON}}")]:
        ws.row_dimensions[row].height = 22
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_RED_BG, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 3. Developer Technical Specification
    # ══════════════════════════════════════════════════
    _section(ws, row, "  3.  Developer Technical Specification")
    row += 1

    # Recommended — full spec
    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, row, 6,
           value="  [ Recommended ]   {{SOL_REC_NAME}}",
           bold=True, size=11, font_color=C_WHITE, fill=C_GREEN_HDR, border=False)
    row += 1

    dev_fields = [
        ("Solution Name / ID", "{{SOL_REC_NAME}}  |  ID: {{SOL_REC_ID}}"),
        ("Rationale",          "{{SOL_REC_TECH_REASON}}"),
        ("Implementation Overview", "{{SOL_REC_IMPL_OVERVIEW}}"),
        ("Prerequisites",      "{{SOL_REC_PREREQ_TECH}}"),
        ("Limitations",        "{{SOL_REC_LIMITS}}"),
        ("Considerations",     "{{SOL_REC_CONSIDERATIONS}}"),
    ]
    for lbl, val in dev_fields:
        h = 50 if lbl in ("Implementation Overview", "Prerequisites",
                          "Limitations", "Considerations") else 22
        ws.row_dimensions[row].height = h
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_GREEN_BG, colspan=5)
        row += 1

    # Risk table
    ws.row_dimensions[row].height = 20
    _label(ws, row, 1, "Risk Details")
    for c in range(2, 7):
        _set(ws, row, c, "")
    row += 1

    for c_idx, h_val in enumerate(["Type", "Severity", "", "", "", ""], 1):
        if h_val:
            _set(ws, row, c_idx, h_val, bold=True, fill=C_BRAND_LIGHT,
                 font_color=C_BRAND_DARK, h="center")
        else:
            _set(ws, row, c_idx, "")
    row += 1

    for r_type, r_level, r_cond, r_act in [
        ("Security",   "High",   "{{RISK_SEC_CONDITION}}", "{{RISK_SEC_ACTION}}"),
        ("License",    "Medium", "{{RISK_LIC_CONDITION}}", "{{RISK_LIC_ACTION}}"),
        ("Operations", "Medium", "{{RISK_OPS_CONDITION}}", "{{RISK_OPS_ACTION}}"),
    ]:
        ws.row_dimensions[row].height = 30
        _set(ws, row, 1, r_type,  bold=True, fill=C_GRAY_BG, h="center")
        _set(ws, row, 2, r_level, fill=C_GRAY_BG, h="center")
        _merge(ws, row, 3, row, 4, value=r_cond, font_color="555555", italic=False, fill=C_WHITE)
        _merge(ws, row, 5, row, 6, value=r_act,  font_color="555555", italic=False, fill=C_WHITE)
        row += 1

    row += 1

    # Needs Review / Not Recommended — compact spec
    for badge_color, badge_text, bg, prefix in [
        (C_YELLOW_HDR, "[ Needs Review ]",    C_YELLOW_BG, "REV"),
        (C_RED_HDR,    "[ Not Recommended ]", C_RED_BG,    "NOT"),
    ]:
        ws.row_dimensions[row].height = 22
        _merge(ws, row, 1, row, 6,
               value=f"  {badge_text}   {{{{SOL_{prefix}_NAME}}}}",
               bold=True, size=11, font_color=C_WHITE, fill=badge_color, border=False)
        row += 1
        for lbl, val in [
            ("Solution Name / ID", f"{{{{SOL_{prefix}_NAME}}}}  |  ID: {{{{SOL_{prefix}_ID}}}}"),
            ("Rationale",          f"{{{{SOL_{prefix}_TECH_REASON}}}}"),
            ("Judgment Basis",     f"{{{{SOL_{prefix}_JUDGMENT}}}}"),
        ]:
            ws.row_dimensions[row].height = 30
            _label(ws, row, 1, lbl)
            _value(ws, row, 2, val, fill=bg, colspan=5)
            row += 1
        row += 1

    # ══════════════════════════════════════════════════
    # 4. AI Deep Analysis
    # ══════════════════════════════════════════════════
    _section(ws, row, "  4.  AI Deep Analysis", C_PURPLE_HDR)
    row += 1

    for lbl, val in [
        ("Analysis Overview",    "{{AI_OVERVIEW}}"),
        ("Key Findings",         "{{AI_FINDINGS}}"),
        ("Optimization Recommendations", "{{AI_RECOMMENDATIONS}}"),
        ("Additional Risks / Opportunities", "{{AI_EXTRA_RISK_OPP}}"),
    ]:
        ws.row_dimensions[row].height = 50
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_PURPLE_BG, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # 5. ROI Estimation
    # ══════════════════════════════════════════════════
    _section(ws, row, "  5.  ROI Estimation", C_TEAL_HDR)
    row += 1

    for lbl, val in [
        ("Current Work Time",     "{{ROI_CURRENT_TIME}}"),
        ("Improved Work Time",    "{{ROI_IMPROVED_TIME}}"),
        ("Time Saved (per unit)", "{{ROI_SAVE_PERIOD}}"),
        ("Annual Time Saved",     "{{ROI_ANNUAL_SAVE}}"),
        ("ROI Payback Period",    "{{ROI_PAYBACK}}"),
        ("Build Time Required",   "{{ROI_BUILD_TIME}}"),
    ]:
        ws.row_dimensions[row].height = 22
        _label(ws, row, 1, lbl)
        _value(ws, row, 2, val, fill=C_TEAL_BG, colspan=5)
        row += 1

    row += 1

    # ══════════════════════════════════════════════════
    # Appendix A. Revision History
    # ══════════════════════════════════════════════════
    _section(ws, row, "  Appendix A.  Revision History", C_GRAY_HDR)
    row += 1

    for c_idx, h_val in enumerate(
        ["Rev", "Type", "Previous Recommendation", "New Recommendation", "Reason for Change", ""], 1
    ):
        _set(ws, row, c_idx, h_val, bold=True, fill=C_BRAND_LIGHT,
             font_color=C_BRAND_DARK, h="center")
    row += 1

    for i in range(1, 4):
        ws.row_dimensions[row].height = 22
        _set(ws, row, 1, f"{{{{REV_{i}_NO}}}}",     fill=C_GRAY_BG, h="center")
        _set(ws, row, 2, f"{{{{REV_{i}_TYPE}}}}",   fill=C_GRAY_BG, h="center")
        _set(ws, row, 3, f"{{{{REV_{i}_PREV}}}}",   font_color="555555", italic=False)
        _set(ws, row, 4, f"{{{{REV_{i}_NEW}}}}",    font_color="555555", italic=False)
        _merge(ws, row, 5, row, 6,
               value=f"{{{{REV_{i}_REASON}}}}",     font_color="555555", italic=False)
        row += 1

    ws.row_dimensions[row].height = 22
    _merge(ws, row, 1, row, 6,
           value="Total Revisions: {{REV_TOTAL}}  (Light: {{REV_LIGHT}}  |  Full: {{REV_FULL}})",
           bold=True, fill=C_BRAND_LIGHT, font_color=C_BRAND_DARK, h="right")
    row += 2

    # ── Footer note ──────────────────────────────────
    _merge(ws, row, 1, row, 6,
           value="Note: Replace all {{VARIABLE_NAME}} cells with actual data before use.",
           italic=False, font_color=C_TEXT_MUTED, fill=C_GRAY_BG,
           h="center", border=False)

    return ws


def main():
    wb = openpyxl.load_workbook(FILE_PATH)
    build_english_sheet(wb)

    # Reorder: KR sheet first, EN sheet second
    kr_sheet = "컨설팅 결과 보고서"
    sheets = wb.sheetnames
    if kr_sheet in sheets and EN_SHEET in sheets:
        kr_idx = sheets.index(kr_sheet)
        en_idx = sheets.index(EN_SHEET)
        if en_idx < kr_idx:
            wb.move_sheet(EN_SHEET, offset=kr_idx - en_idx)

    wb.save(FILE_PATH)
    print(f"[OK] English sheet added: {FILE_PATH}")
    print(f"     Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
